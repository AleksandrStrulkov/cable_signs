import os
import sys
import shutil
import logging
from datetime import datetime


# ДЕТАЛЬНАЯ ДИАГНОСТИКА
if hasattr(sys, '_MEIPASS'):
    base_path = sys._MEIPASS
    print(f"Base path: {base_path}")

    # Проверим все возможные расположения
    possible_paths = [
            os.path.join(base_path, 'tk', 'tk8.6'),
            os.path.join(base_path, 'tcl', 'tk8.6'),
            os.path.join(base_path, 'tk8.6'),
    ]

    for path in possible_paths:
        exists = os.path.exists(path)
        print(f"Path: {path} - exists: {exists}")
        if exists:
            files = os.listdir(path)
            print(f"  Files: {len(files)}")
            if 'tk.tcl' in files:
                print("  ✅ tk.tcl FOUND!")

    # Если tk8.6 в корне как tk/tk8.6, перемещаем
    tk_root_path = os.path.join(base_path, 'tk', 'tk8.6')
    tk_correct_path = os.path.join(base_path, 'tcl', 'tk8.6')

    if os.path.exists(tk_root_path):
        print(f"Found tk8.6 at: {tk_root_path}")
        if not os.path.exists(tk_correct_path):
            os.makedirs(os.path.dirname(tk_correct_path), exist_ok=True)
            shutil.copytree(tk_root_path, tk_correct_path)
            print(f"✅ Copied tk8.6 to: {tk_correct_path}")

    # Устанавливаем пути
    os.environ['TCL_LIBRARY'] = os.path.join(base_path, 'tcl', 'tcl8.6')
    os.environ['TK_LIBRARY'] = os.path.join(base_path, 'tcl', 'tk8.6')

    # print(f"Final TCL_LIBRARY: {os.environ['TCL_LIBRARY']} - exists: {os.path.exists(os.environ['TCL_LIBRARY'])}")
    # print(f"Final TK_LIBRARY: {os.environ['TK_LIBRARY']} - exists: {os.path.exists(os.environ['TK_LIBRARY'])}")

import tkinter as tk

# with open('debug_log.txt', 'a') as f:
#     f.write("Tkinter imported successfully\n")

# import tkinter as tk
import openpyxl
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
# Для кастомной темы
from tkinter import ttk, filedialog, messagebox

# Настройка логирования в файл И в консоль
log_filename = f"cable_signs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"


class ColoredFormatter(logging.Formatter):
    """Добавляет цвета в консольные логи"""
    COLORS = {
            'INFO': '\033[94m',  # Синий
            'WARNING': '\033[93m',  # Желтый
            'ERROR': '\033[91m',  # Красный
            'CRITICAL': '\033[91m',  # Красный
            'RESET': '\033[0m'  # Сброс
    }

    def format(self, record):
        log_message = super().format(record)
        if record.levelname in self.COLORS:
            return f"{self.COLORS[record.levelname]}{log_message}{self.COLORS['RESET']}"
        return log_message


# Создаем логгер
logger = logging.getLogger()
logger.setLevel(logging.INFO)

# Цветной формат для консоли
colored_formatter = ColoredFormatter('%(levelname)s: %(message)s')

# Обычный формат для файла
file_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')

# Файловый обработчик
file_handler = logging.FileHandler('app.log', encoding='utf-8')
file_handler.setFormatter(file_formatter)

# Консольный обработчик
console_handler = logging.StreamHandler(sys.stdout)
console_handler.setFormatter(colored_formatter)

# Добавляем оба обработчика
logger.addHandler(file_handler)
logger.addHandler(console_handler)

# Логируем запуск
logging.info("=== Cable Signs Application Started ===")
logging.info(f"Working directory: {os.getcwd()}")

# Теперь все логи будут и в файл, и в консоль
logging.info("This message goes to both file and console")
logging.warning("This warning is visible in console")
logging.error("Errors also appear in console")

# Указываем Python, где искать Tcl/Tk внутри виртуального окружения
base_prefix = getattr(sys, 'base_prefix', sys.prefix)  # Получаем путь к окружению
tcl_dir = os.path.join(base_prefix, 'tcl')

# Важно: задаем переменные окружения ДО создания окна Tk
os.environ["TCL_LIBRARY"] = os.path.join(tcl_dir, "tcl8.6")
os.environ["TK_LIBRARY"] = os.path.join(tcl_dir, "tk8.6")

# === РЕГИСТРАЦИЯ ШРИФТА ===
# Попробуем загрузить Times New Roman Bold для красивого шрифта
try:
    pdfmetrics.registerFont(TTFont('Times-Bold', 'timesbd.ttf'))
except:
    pass  # Если не найден — используем fallback (стандартный жирный шрифт)

# === ГЛОБАЛЬНЫЕ ПАРАМЕТРЫ ===
TRIANGLE_BASE = 60 * mm        # Ширина основания — 60 мм (по ГОСТ)
TRIANGLE_HEIGHT = 55 * mm      # Высота — 55 мм (по ГОСТ)
PAGE_WIDTH, PAGE_HEIGHT = A4  # Размер листа A4

MAX_COLS = 5  # Количество треугольников в ряду
MAX_ROWS = 5  # Количество рядов

FONT_SYSTEM = 24  # Размер шрифта для system
FONT_TRACK = 14  # Размер шрифта для track
FONT_CABLE = 16  # Размер шрифта для cable (чуть больше)
FONT_LENGTH = 14  # Размер шрифта для length

MIN_FONT_SIZE = 10  # Минимальный размер шрифта при уменьшении
PRINTER_OFFSET_X = 0.0 * mm  # Компенсация смещения принтера на обратной стороне по оси X
PRINTER_OFFSET_Y = 0.0 * mm  # Компенсация смещения принтера на обратной стороне по оси Y
# Запрещённые символы в именах файлов Windows
INVALID_FILENAME_CHARS = r'<>:"/\\|?*'


# === ФУНКЦИЯ: РАЗДЕЛЕНИЕ ТЕКСТА CABLE НА 2 СТРОКИ ===
def split_cable_text(text):
    """
    Разделяет текст из колонки 'cable' на 2 строки:
    - Первая строка: первое слово
    - Вторая строка: всё остальное
    Если одно слово — делит пополам.
    Пример: "ParLan 4x2x0,57" → ["ParLan", "4x2x0,57"]
    """
    if not text or not text.strip():
        return ["", ""]
    words = text.strip().split()
    if len(words) == 0:
        return ["", ""]
    elif len(words) == 1:
        w = words[0]
        mid = len(w) // 2
        return [w[:mid], w[mid:]]
    else:
        return [words[0], " ".join(words[1:])]


def find_column(headers, *names):
    """
    Ищет первый столбец по списку возможных имён.
    :param headers: список заголовков Excel
    :param names: возможные названия столбца
    :return: индекс столбца или None
    """
    lower_headers = [h.lower() if h else "" for h in headers]

    for name in names:
        if name.lower() in lower_headers:
            return lower_headers.index(name.lower())

    return None


# === ОСНОВНОЙ КЛАСС ПРИЛОЖЕНИЯ ===
class CableLabelApp:
    def __init__(self, root):
        """
        Инициализация GUI приложения.
        :param root: главное окно Tkinter
        """
        # Переменные для компенсации принтера (в мм)
        self.printer_offset_x = tk.StringVar(value="0.0")
        self.printer_offset_y = tk.StringVar(value="0.0")
        # Внутренние float-значения
        self._offset_x = 0.0
        self._offset_y = 0.0
        self.root = root
        # --- Настройка тёмной темы ---
        self.root.tk_setPalette(
                background='#2e2e2e', foreground='white',
                activeBackground='#4a4a4a', activeForeground='white'
        )
        # Переменная для толщины контура
        self.line_width_var = tk.StringVar(value="5.0")
        self._line_width = 5.0  # внутреннее значение в мм

        # 🔔 Подписываемся на изменение
        self.line_width_var.trace_add('write', self.update_offsets)

        style = ttk.Style()
        style.theme_use('clam')  # или 'alt'

        style.configure('.', background='#2e2e2e', foreground='white', fieldbackground='#3c3c3c')
        style.configure('TLabel', foreground='white', background='#2e2e2e')
        style.configure('TButton', background='#007acc', foreground='white', padding=5)
        style.map('TButton', background=[('active', '#005a99')])
        style.configure('TEntry', fieldbackground='#3c3c3c', foreground='white', insertcolor='white')
        style.configure('TProgressbar', background='#007acc', troughcolor='#1e1e1e')

        # Цвет текста справки — светло-серый
        self.help_color = "#ccccff"
        self.root.title('Генератор бирок')
        self.root.geometry("580x630")

        self.input_file = tk.StringVar()  # Путь к Excel
        self.output_dir = tk.StringVar()  # Путь к папке сохранения

        self.create_widgets()

    def sanitize_filename(self, name):
        """Заменяет запрещённые символы на _"""
        for char in INVALID_FILENAME_CHARS:
            name = name.replace(char, '_')
            logger.info(f"🧹 Заменены запрещенные символы в имени файла: {char} на {name}")
        return name.strip()

    def reset_filename(self):
        """Сбросить имя файла на значение по умолчанию"""
        self.output_name.set("cable_labels.pdf")

    def create_widgets(self):
        """Создаёт элементы интерфейса"""
        frame = ttk.Frame(self.root, padding=20)
        frame.pack(fill="both", expand=True)

        # Регистрируем функцию валидации
        validate_cmd = (self.root.register(self.validate_float_input), '%P')

        # Заголовок
        ttk.Label(frame, text="Генератор бирок под маркировку трасс кабеля", font=("Arial", 14, "bold")).grid(
                row=0, column=0, columnspan=3, pady=(0, 15)
        )

        # Excel файл
        ttk.Label(frame, text="Excel файл:").grid(row=1, column=0, sticky="w", pady=5)
        ttk.Entry(frame, textvariable=self.input_file, width=40).grid(row=1, column=1, padx=5, pady=5)
        ttk.Button(frame, text="Обзор", command=self.browse_input).grid(row=1, column=2, padx=5)

        # Папка сохранения
        ttk.Label(frame, text="Папка сохранения:").grid(row=2, column=0, sticky="w", pady=5)
        ttk.Entry(frame, textvariable=self.output_dir, width=40).grid(row=2, column=1, padx=5, pady=5)
        ttk.Button(frame, text="Обзор", command=self.browse_output).grid(row=2, column=2, padx=5)

        # Имя выходного файла
        ttk.Label(frame, text="Имя PDF-файла:").grid(row=3, column=0, sticky="w", pady=5)
        self.output_name = tk.StringVar(value="cable_labels")  # значение по умолчанию
        # ttk.Button(frame, text="По умолчанию", command=self.reset_filename).grid(row=3, column=3, padx=5)
        ttk.Entry(frame, textvariable=self.output_name, width=40).grid(row=3, column=1, padx=5, pady=5)
        ttk.Label(frame, text=".pdf", foreground="gray").grid(row=3, column=2, sticky="w", padx=(0, 5))

        # Справка
        help_text = (
                "1. Создайте файл Excel с колонками:\n"
                "_______________________________________\n"
                "| Подсистема | Трасса | Кабель | Длина | Кол-во |\n"
                "¯¯¯¯¯¯¯¯¯¯¯¯¯¯¯¯¯¯¯¯¯¯¯¯¯¯¯¯¯¯¯¯¯¯¯¯¯¯¯\n"
                "Где 'Кол-во' — количество бирок на трассу.\n\n"
                "2. Загрузите файл Excel и выберите папку для сохранения.\n\n"
                "2. Укажите имя выходного файла.\n\n"
                "3. Нажмите 'Создать PDF'."
        )
        help_label = ttk.Label(
                frame,
                text=help_text,
                font=("Arial", 9),
                foreground=self.help_color,
                relief="flat",
                anchor="center",
                justify="left",
                wraplength=460
        )
        help_label.grid(row=4, column=0, columnspan=3, pady=(15, 10), sticky="ew")

        # Толщина контура
        ttk.Label(frame, text="Толщина контура (мм):").grid(row=5, column=0, sticky="w", pady=5)
        ttk.Entry(
                frame,
                textvariable=self.line_width_var,
                width=8,
                validate='key',
                validatecommand=(self.root.register(self.validate_float_input), '%P')
        ).grid(row=5, column=1, sticky="w", padx=(0, 10))

        # Подсказка
        width_hint = ttk.Label(
                frame,
                text="Рекомендуется: 1.8–6.0 мм",
                font=("Arial", 8),
                foreground="gray"
        )
        width_hint.grid(row=6, column=0, columnspan=4, sticky="w", pady=(0, 5))

        # --- Компенсация принтера ---
        ttk.Label(frame, text="Компенсация печати по оси X (мм):").grid(row=7, column=0, sticky="w", pady=(15, 5))
        ttk.Entry(
                frame,
                textvariable=self.printer_offset_x,
                width=8,
                validate='key',
                validatecommand=validate_cmd
        ).grid(row=7, column=1, sticky="w", padx=(0, 10))

        ttk.Label(frame, text="Компенсация печати по оси Y (мм):").grid(row=8, column=0, sticky="w", pady=5)
        ttk.Entry(
                frame,
                textvariable=self.printer_offset_y,
                width=8,
                validate='key',
                validatecommand=validate_cmd
        ).grid(row=8, column=1, sticky="w", padx=(0, 10))

        # Подпишемся на изменения
        self.printer_offset_x.trace_add('write', self.update_offsets)
        self.printer_offset_y.trace_add('write', self.update_offsets)

        # Подсказка
        offset_hint = ttk.Label(
                frame,
                text="Смещение применяется только на обратной стороне\nИспользуйте переплет по длинному краю ",
                font=("Arial", 8),
                foreground="gray"
        )
        offset_hint.grid(row=9, column=0, columnspan=4, sticky="w", pady=(0, 10))

        # Кнопка генерации
        ttk.Button(frame, text="Создать PDF", command=self.generate).grid(row=10, column=0, columnspan=3, pady=10)

        # Прогресс бар
        self.progress = ttk.Progressbar(frame, mode="determinate")
        self.progress.grid(row=11, column=0, columnspan=3, pady=10, sticky="ew")

        # Подпись компании — в левый нижний угол
        copyright_label = tk.Label(
                frame,
                text='@2025 ООО "ДГС" УГПР №2',
                font=("Arial", 7),
                fg="#468000",  # Тёмно-зелёный цвет
                bg="#2e2e2e",  # Совпадает с фоном (для тёмной темы)
                anchor="w"
        )
        copyright_label.grid(row=12, column=0, columnspan=2, sticky="w", pady=(10, 0))

    def validate_float_input(self, value_if_allowed):
        """
        Разрешает ввод чисел, знака и десятичного разделителя (., ,)
        """
        if value_if_allowed == "" or value_if_allowed in ["-", ".", ","]:
            return True
        try:
            value_if_allowed.replace(',', '.').replace('-', '').strip()
            float(value_if_allowed.replace(',', '.'))
            return True
        except ValueError:
            return False

    def update_offsets(self, *args):
        """Обновляет внутренние float-значения всех параметров"""
        try:
            x_val = self.printer_offset_x.get().strip()
            self._offset_x = float(x_val.replace(',', '.')) if x_val not in ('', '-', '.') else 0.0
        except:
            self._offset_x = 0.0

        try:
            y_val = self.printer_offset_y.get().strip()
            self._offset_y = float(y_val.replace(',', '.')) if y_val not in ('', '-', '.') else 0.0
        except:
            self._offset_y = 0.0

        try:
            w_val = self.line_width_var.get().strip()
            self._line_width = float(w_val.replace(',', '.')) if w_val not in ('', '-', '.') else 5.0
            if self._line_width <= 0:
                self._line_width = 0.1
        except:
            self._line_width = 5.0

    def browse_input(self):
        """Выбор Excel файла"""
        file = filedialog.askopenfilename(title="Выберите Excel", filetypes=[("Excel", "*.xlsx")])
        if file:
            self.input_file.set(file)
            # Получаем информацию о файле
            file_size = os.path.getsize(file)  # Размер в байтах
            file_time = datetime.fromtimestamp(
                    os.path.getmtime(file)
            ).strftime('%Y-%m-%d %H:%M:%S')  # Время изменения

            logging.info(f"📁 Файл загружен успешно с пути: {file}")
            logging.info(f"📊 Размер: {file_size} байт")
            logging.info(f"🕒 Изменен: {file_time}")

    def browse_output(self):
        """Выбор папки для сохранения"""
        folder = filedialog.askdirectory(title="Выберите папку")
        if folder:
            self.output_dir.set(folder)
            # Получаем информацию о файле
            logger.info(f"📁 Выбрана папка для сохранения по пути: {folder}")

    def generate(self):
        """Основная логика генерации PDF"""
        input_path = self.input_file.get()
        output_dir = self.output_dir.get()

        if not input_path or not output_dir:
            messagebox.showerror("Ошибка", "Укажите файл и папку!")
            logger.error(f"🚨 Ошибка: Не указан файл или папка для сохранения.\n")
            return

        try:
            wb = openpyxl.load_workbook(input_path)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]

            # Поиск индексов столбцов
            system_idx = find_column(headers, "system", "Подсистема", "Система")
            track_idx = find_column(headers, "track", "Трасса", "Обозначение")
            cable_idx = find_column(headers, "cable", "Кабель")
            length_idx = find_column(headers, "length", "Длина")
            quantity_idx = find_column(headers, "quantity", "Количество", "Кол-во")
            list_idx = [system_idx, track_idx, cable_idx, length_idx, quantity_idx]

            headers_not_none = [item for item in headers if item is not None]
            border = '-'
            border_headers_not_none = 0 + 28
            for header in headers_not_none:
                len_header = len(header)
                border_headers_not_none += len_header

            logging.info(
                f"🚀 Загруженный файл excel имеет заголовки:\n"
                f"{border * border_headers_not_none}\n"
                f"| {headers_not_none} |\n"
                f"{border * border_headers_not_none}\n"
                )

            if None in (system_idx, track_idx, cable_idx, length_idx, quantity_idx):
                if None in list_idx:
                    for i, idx in enumerate(list_idx):
                        if idx is None:
                            list_idx[i] = f"{i + 1} (не найден)"
                        elif idx is not None:
                            list_idx[i] = f"{i + 1} ({headers[idx]})"
                messagebox.showerror(
                        "Ошибка", "Не найдены необходимые столбцы!"
                                  "\nПроверьте Excel файл и повторите попытку.\n"
                                  f"\nНайденные столбцы:\n {list_idx}\n"
                                  f"Не забудьте сохранить файл после редактирования!"
                )
                logging.error(f"🚨 Ошибка: Не найдены необходимые столбцы!\n")

                return

            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                system = str(row[system_idx] or "").strip()
                track = str(row[track_idx] or "").strip()
                cable = str(row[cable_idx] or "").strip()
                length_val = str(row[length_idx] or "").strip()
                try:
                    qty = int(row[quantity_idx])
                except:
                    qty = 1
                for _ in range(qty):
                    data.append(
                            {
                                    "system": system,
                                    "track": track,
                                    "cable": cable,
                                    "length": length_val
                            }
                    )

            # Получаем имя файла от пользователя
            file_name = self.output_name.get().strip()
            if not file_name.endswith(".pdf"):
                file_name += ".pdf"

            # Получаем и очищаем имя файла
            raw_name = self.output_name.get().strip()
            if not raw_name:
                messagebox.showwarning("Предупреждение", "Введите имя файла.")
                logger.error(f"🚨 Ошибка: Имя файла пустое.")
                return

            clean_name = self.sanitize_filename(raw_name)
            logger.info(f"📝 Задано имя выходного файла: {clean_name + '.pdf'}")

            if not clean_name:
                messagebox.showerror("Ошибка", "Имя файла пустое после очистки.")
                logging.error(f"🚨 Ошибка: Имя файла пустое после очистки.")
                return

            if not clean_name.endswith(".pdf"):
                clean_name += ".pdf"

            output_path = os.path.join(output_dir, clean_name)
            # Заменяем обратные слеши на прямые
            normalized_path = output_path.replace('\\', '/')
            logger.info(f"📝 Выходной файл pdf сохранен по пути: {normalized_path}")

            # Проверяем, можно ли создать файл
            try:
                with open(output_path, 'w'):
                    pass
                os.remove(output_path)  # чистим тестовый файл
            except Exception as e:
                messagebox.showerror("Ошибка", f"Невозможно создать файл:\n{clean_name}\n\n{str(e)}")
                logger.error(f"🚨 Ошибка при создании файла: {str(e)}")
                return

            c = canvas.Canvas(output_path, pagesize=A4)
            c.setFont("Times-Bold", 12)

            index = 0
            total_sides = len(data) * 2
            self.progress["maximum"] = total_sides
            self.progress["value"] = 0

            while index < len(data):
                # Лицевая сторона
                self.draw_page(c, data, index, side='front')
                c.showPage()
                self.progress["value"] += 1

                # Обратная сторона
                self.draw_page(c, data, index, side='back')
                if index + MAX_COLS * MAX_ROWS < len(data):
                    c.showPage()
                self.progress["value"] += 1

                index += MAX_COLS * MAX_ROWS

            c.save()
            messagebox.showinfo("Готово", f"PDF сохранён:\n{output_path}")
            logger.info(f"📝 Выходной файл pdf сохранен по пути: {normalized_path}")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка: {str(e)}")
            logger.error(f"🚨 Ошибка: {str(e)}")

    def draw_page(self, c, data, start_index, side):
        """
        Рисует одну страницу с этикетками.
        :param c: объект canvas (PDF)
        :param data: список данных
        :param start_index: с какого элемента начинать
        :param side: 'front' или 'back'
        """
        col_step = TRIANGLE_BASE / 2
        x_centers_original = [45 * mm, 75 * mm, 105 * mm, 135 * mm, 165 * mm]
        Y_START = 70 * mm # Начальная Y координата первого ряда

        # Компенсация принтера — только на обратной стороне
        shift_x = self._offset_x * mm if side == 'back' else 0
        shift_y = self._offset_y * mm if side == 'back' else 0

        # Для обратной стороны — отзеркаливаем X координаты
        count = 0
        for i in range(start_index, min(start_index + MAX_COLS * MAX_ROWS, len(data))):
            item = data[i]
            col = count % MAX_COLS
            row = count // MAX_COLS

            if row >= MAX_ROWS:
                break

            # Базовая координата X (для лицевой стороны)
            center_x_base = x_centers_original[col]

            # Для обратной стороны — отзеркаливаем относительно центра листа
            if side == 'back':
                center_x = PAGE_WIDTH - center_x_base + shift_x
                # center_x = x_centers_original[col] + shift_x
            else:
                center_x = center_x_base + shift_x

            # y_base = Y_START + row * TRIANGLE_HEIGHT
            y_base = Y_START + row * TRIANGLE_HEIGHT
            is_upside_down = col % 2 == 1

            if side == 'front':
                main_text = item["system"]
                sub_text = item["track"]
                main_font = FONT_SYSTEM
                sub_font = FONT_TRACK
            else:
                main_text = item["cable"]
                raw_length = str(item["length"]).strip()
                sub_text = f"L={raw_length} м" if raw_length.replace('.', '').isdigit() else raw_length
                main_font = FONT_CABLE
                sub_font = FONT_LENGTH

            self.draw_triangle(
                    c, center_x, y_base, is_upside_down, main_text, sub_text,
                    main_font, sub_font, side
            )

            count += 1

    def draw_triangle(self, c, center_x, y_base, upside_down, main_text, sub_text,
                      main_font_size, sub_font_size, side):
        """
        Рисует один треугольник 60×55 мм по ГОСТ.
        :param c: canvas
        :param center_x: X центра основания
        :param y_base: Y основания
        :param upside_down: True если остриём вниз
        :param main_text: основной текст
        :param sub_text: подзаголовок
        :param main_font_size: размер шрифта основного текста
        :param sub_font_size: размер шрифта подзаголовка
        :param side: 'front' или 'back'
        """
        base = TRIANGLE_BASE
        height = TRIANGLE_HEIGHT
        x_left = center_x - base / 2
        x_right = center_x + base / 2

        # Вершины треугольника (по ГОСТ: основание 60 мм, высота 55 мм)
        if upside_down:
            points = [(x_left, y_base), (x_right, y_base), (center_x, y_base - height)]
        else:
            points = [(x_left, y_base - height), (x_right, y_base - height), (center_x, y_base)]

        # Контур
        c.setLineWidth(self._line_width)
        c.setStrokeColorRGB(0, 0, 0)
        c.lines([
            (points[0][0], points[0][1], points[1][0], points[1][1]),
            (points[1][0], points[1][1], points[2][0], points[2][1]),
            (points[2][0], points[2][1], points[0][0], points[0][1])
        ])

        # Относительные позиции (в долях от высоты)
        dy_main = height * 0.38  # Основной текст — чуть выше центра
        dy_sub = height * 0.1   # Подзаголовок — у основания

        c.saveState()

        if upside_down:
            # Поворачиваем вокруг основания
            c.translate(center_x, y_base)
            c.rotate(180)
            c.translate(-center_x, -y_base)
            y_main = y_base + dy_main
            y_sub = y_base + dy_sub
        else:
            base_y = y_base - height
            y_main = base_y + dy_main
            y_sub = base_y + dy_sub

        # --- ОСНОВНОЙ ТЕКСТ ---
        if side == 'back':  # Это обратная сторона — cable
            parts = split_cable_text(main_text)
            line1, line2 = parts[0], parts[1]

            # Шрифт для первой строки — всегда FONT_CABLE (16)
            fs_line1 = FONT_CABLE

            # Шрифт для второй строки — уменьшаем, если длинная
            if len(line2) >= 15:
                fs_line2 = 12
            elif len(line2) >= 10:
                fs_line2 = 14
            else:
                fs_line2 = FONT_CABLE  # 16

            # Позиции строк: первая — чуть выше центра, вторая — чуть ниже
            y_upper = y_main - fs_line1 * 0.5  # выше
            y_lower = y_main + fs_line2 * 0.5  # ниже
            y_positions = [y_lower, y_upper]  # ⚠️ Важно: сначала верхняя, потом нижняя

            # Рисуем первую строку
            c.setFont("Times-Bold", fs_line1)
            try:
                tw1 = pdfmetrics.stringWidth(line1, "Times-Bold", fs_line1)
            except:
                tw1 = len(line1) * fs_line1 * 0.6
            x_pos1 = center_x - tw1 / 2
            c.drawString(x_pos1, y_positions[0], line1)  # Первая строка — выше

            # Рисуем вторую строку
            c.setFont("Times-Bold", fs_line2)
            try:
                tw2 = pdfmetrics.stringWidth(line2, "Times-Bold", fs_line2)
            except:
                tw2 = len(line2) * fs_line2 * 0.6
            x_pos2 = center_x - tw2 / 2
            c.drawString(x_pos2, y_positions[1], line2)  # Вторая строка — ниже

        else:  # Лицевая сторона — system
            lines = [main_text]

            # Плавное уменьшение от 22 до 16 pt при росте длины
            length = len(main_text)
            fs = max(16, 32 - length * 2) if length >= 5 else FONT_SYSTEM


            if fs < MIN_FONT_SIZE:
                fs = MIN_FONT_SIZE

            y_positions = [y_main]

            c.setFont("Times-Bold", fs)
            for j, line in enumerate(lines):
                try:
                    tw = pdfmetrics.stringWidth(line, "Times-Bold", fs)
                except:
                    tw = len(line) * fs * 0.6
                x_pos = center_x - tw / 2
                y_pos = y_positions[j]
                c.drawString(x_pos, y_pos, line)

        # --- ПОДЗАГОЛОВОК (track или length) ---
        base_font_size = sub_font_size  # 14 pt
        line_spacing = base_font_size * 1.1

        if side == 'front':
            try:
                full_width = pdfmetrics.stringWidth(sub_text, "Times-Bold", base_font_size)
                max_width = TRIANGLE_BASE * 0.9
                fits_in_one_line = full_width <= max_width
            except:
                fits_in_one_line = len(sub_text) <= 30

            if not fits_in_one_line and '/' in sub_text:
                parts = sub_text.split('/', 1)
                line1 = parts[0] + '/'
                line2 = parts[1].strip()
                lines = [line1, line2]

                # Определяем максимальную длину среди двух частей

                max_str = max(parts[0], parts[1])
                max_ind = parts.index(max_str)
                max_len = max(len(parts[0]), len(parts[1]))               


                # Определяем максимальную длину среди двух частей
                len1, len2 = len(parts[0]), len(parts[1])
                max_len = max(len1, len2)
                max_ind = 0 if len1 >= len2 else 1

                # Плавное уменьшение шрифта для ОБЕИХ строк
                if max_len < 15:
                    track_font_size = 14.0
                elif max_len == 15:
                    track_font_size = 13.0
                elif max_ind == 1 and max_len >= 18:
                    track_font_size = 10.0  # вторая часть длинная — сильно уменьшаем
                elif max_ind == 0 and max_len >= 18:
                    track_font_size = 12.0  # первая часть длинная — умеренно уменьшаем
                elif max_ind == 0:
                    track_font_size = max(11.0, 14.0 - (max_len - 15) * 1.2)
                else:  # max_ind == 1 и max_len < 18
                    track_font_size = max(10.0, 14.0 - (max_len - 15) * 0.5)

            else:
                lines = [sub_text]
                track_font_size = base_font_size  # 14 pt

        else:
            lines = [sub_text]
            track_font_size = base_font_size

        lines = lines[:2]

        # Базовая Y: где должна быть одна строка
        y_single = y_sub + height * 0.05

        if len(lines) == 2:
            y_pos_1 = y_single - line_spacing
            y_pos_2 = y_single
            y_positions = [y_pos_2, y_pos_1]  # первая выше, вторая ниже
        else:
            y_positions = [y_single]

        # Рисуем все строки одним шрифтом
        c.setFont("Times-Bold", track_font_size)

        for j, line in enumerate(lines):
            if not line.strip():
                continue
            try:
                tw = pdfmetrics.stringWidth(line, "Times-Bold", track_font_size)
            except:
                estimated_width_per_char = {
                    'Ш': 1.2, 'Щ': 1.2, 'Ж': 1.15, 'Д': 1.1, 'П': 1.05,
                    'А': 0.9, 'В': 0.95, 'Е': 0.9, 'К': 0.95, 'Х': 0.9
                }
                total_width = sum(estimated_width_per_char.get(c.upper(), 1.0) for c in line)
                tw = total_width * track_font_size * 0.58

            x_pos = center_x - tw / 2
            y_pos = y_positions[j]
            c.drawString(x_pos, y_pos, line)

        c.restoreState()


if __name__ == "__main__":
    root = tk.Tk()
    app = CableLabelApp(root)
    root.mainloop()