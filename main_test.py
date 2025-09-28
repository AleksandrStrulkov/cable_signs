import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Указываем Python, где искать Tcl/Tk внутри виртуального окружения
base_prefix = getattr(sys, 'base_prefix', sys.prefix)  # Получаем путь к окружению
tcl_dir = os.path.join(base_prefix, 'tcl')

# Важно: задаем переменные окружения ДО создания окна Tk
os.environ["TCL_LIBRARY"] = os.path.join(tcl_dir, "tcl8.6")
os.environ["TK_LIBRARY"] = os.path.join(tcl_dir, "tk8.6")

# === РЕГИСТРАЦИЯ ШРИФТА ===
# Попробуем загрузить Times New Roman Bold для красивого шрифта
try:
    pdfmetrics.registerFont(TTFont('Times-Bold', 'timesbd.ttf'))
    print("✅ Шрифт Times-Bold успешно загружен")
except:
    pass  # Если не найден — используем fallback (стандартный жирный шрифт)

# === ГЛОБАЛЬНЫЕ ПАРАМЕТРЫ ===
TRIANGLE_BASE = 60 * mm  # Ширина основания треугольника
TRIANGLE_HEIGHT = 49 * mm  # Высота треугольника
PAGE_WIDTH, PAGE_HEIGHT = A4  # Размер листа A4

MAX_COLS = 5  # Количество треугольников в ряду
MAX_ROWS = 5  # Количество рядов

FONT_SYSTEM = 18  # Размер шрифта для system
FONT_TRACK = 14  # Размер шрифта для track
FONT_CABLE = 16  # Размер шрифта для cable (чуть больше)
FONT_LENGTH = 14  # Размер шрифта для length

MIN_FONT_SIZE = 10  # Минимальный размер шрифта при уменьшении
PRINTER_OFFSET_X = 0.0 * mm  # Компенсация смещения принтера на обратной стороне


# === ФУНКЦИЯ: РАЗДЕЛЕНИЕ ТЕКСТА CABLE НА 2 СТРОКИ ===
def split_cable_text(text):
    """
    Разделяет текст из колонки 'cable' на 2 строки:
    - Первая строка: первое слово
    - Вторая строка: всё остальное
    Если одно слово — делит пополам.
    Пример: "ParLan 4x2x0,57" → ["ParLan", "4x2x0,57"]
    """
    if not text or not text.strip():
        return ["", ""]
    words = text.strip().split()
    if len(words) == 0:
        return ["", ""]
    elif len(words) == 1:
        w = words[0]
        mid = len(w) // 2
        return [w[:mid], w[mid:]]
    else:
        return [words[0], " ".join(words[1:])]


# === ФУНКЦИЯ: РАЗДЕЛЕНИЕ ТЕКСТА CABLE НА 2 СТРОКИ ===
# def split_track_text(text):
#     """
#     Разделяет текст из колонки 'track' на 2 строки:
#     - Первая строка: первое слово
#     - Вторая строка: всё остальное
#     Если одно слово — делит пополам.
#     Пример: "1ШСУ14-БП1/ТШМ-60.01" → ["1ШСУ14-БП1", "ТШМ-60.01"]
#     """
#     if not text or not text.strip():
#         return ["", ""]
#     words = text.strip().split("/")
#     if len(words) == 0:
#         return ["", ""]
#     elif len(words) == 1:
#         w = words[0]
#         mid = len(w) // 2
#         return [w[:mid], w[mid:]]
#     else:
#         return [words[0] + "/", " ".join(words[1:])]


# print(split_track_text("1ШСУ14-БП1hhfgxxhrte/sfgnfxftТШМ-60.01"))


def find_column(headers, *names):
    """
    Ищет первый столбец по списку возможных имён.
    :param headers: список заголовков Excel
    :param names: возможные названия столбца
    :return: индекс столбца или None
    """
    lower_headers = [h.lower() if h else "" for h in headers]
    for name in names:
        if name.lower() in lower_headers:
            return lower_headers.index(name.lower())
    return None


# === ОСНОВНОЙ КЛАСС ПРИЛОЖЕНИЯ ===
class CableLabelApp:
    def __init__(self, root):
        """
        Инициализация GUI приложения.
        :param root: главное окно Tkinter
        """
        self.root = root
        self.root.title("Генератор бирок")
        self.root.geometry("500x300")

        self.input_file = tk.StringVar()  # Путь к Excel
        self.output_dir = tk.StringVar()  # Путь к папке сохранения

        self.create_widgets()

    def create_widgets(self):
        """Создаёт элементы интерфейса"""
        frame = ttk.Frame(self.root, padding=20)
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text="Генератор бирок под маркировку трасс кабеля", font=("Arial", 14, "bold")).grid(
                row=0, column=0, columnspan=3, pady=10
        )
        ttk.Label(
                frame, text="1. Создайте файл excel с колонками:\n"
                            "_______________________________________\n"
                            "|Подсистема|Трасса|Кабель|Длина|Кол-во|\n "
                            "_______________________________________\n"
                            "Где 'кол-во' — количество бирок на трассу кабеля.\n"
                            "2. Загрузите файл excel и выберите папку для сохранения PDF.\n"
                            "3. Нажмите 'Создать PDF'.",
                font=("Arial", 10, "bold")
        ).grid(row=0, column=0, columnspan=3, pady=10)
        ttk.Label(frame, text="Excel файл:").grid(row=1, column=0, sticky="w", pady=5)
        ttk.Entry(frame, textvariable=self.input_file, width=40).grid(row=1, column=1, padx=5, pady=5)
        ttk.Button(frame, text="Обзор", command=self.browse_input).grid(row=1, column=2, padx=5)

        ttk.Label(frame, text="Папка сохранения:").grid(row=2, column=0, sticky="w", pady=5)
        ttk.Entry(frame, textvariable=self.output_dir, width=40).grid(row=2, column=1, padx=5, pady=5)
        ttk.Button(frame, text="Обзор", command=self.browse_output).grid(row=2, column=2, padx=5)

        ttk.Button(frame, text="Создать PDF", command=self.generate).grid(row=3, column=0, columnspan=3, pady=20)

        self.progress = ttk.Progressbar(frame, mode="determinate")
        self.progress.grid(row=4, column=0, columnspan=3, pady=10, sticky="ew")

    def browse_input(self):
        """Выбор Excel файла"""
        file = filedialog.askopenfilename(title="Выберите Excel", filetypes=[("Excel", "*.xlsx")])
        if file:
            self.input_file.set(file)

    def browse_output(self):
        """Выбор папки для сохранения"""
        folder = filedialog.askdirectory(title="Выберите папку")
        if folder:
            self.output_dir.set(folder)

    def generate(self):
        """Основная логика генерации PDF"""
        input_path = self.input_file.get()
        output_dir = self.output_dir.get()

        if not input_path or not output_dir:
            messagebox.showerror("Ошибка", "Укажите файл и папку!")
            return

        try:
            wb = openpyxl.load_workbook(input_path)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]

            # Поиск индексов столбцов
            system_idx = find_column(headers, "system", "Подсистема", "Система")
            track_idx = find_column(headers, "track", "Трасса", "Обозначение")
            cable_idx = find_column(headers, "cable", "Кабель")
            length_idx = find_column(headers, "length", "Длина")
            quantity_idx = find_column(headers, "quantity", "Количество", "Кол-во")
            list_idx = [system_idx, track_idx, cable_idx, length_idx, quantity_idx]

            if None in (system_idx, track_idx, cable_idx, length_idx, quantity_idx):
                if None in list_idx:
                    for i, idx in enumerate(list_idx):
                        if idx is None:
                            list_idx[i] = f"{i + 1} (не найден)"
                        elif idx is not None:
                            list_idx[i] = f"{i + 1} ({headers[idx]})"
                messagebox.showerror(
                        "Ошибка", "Не найдены необходимые столбцы!"
                                  "\nПроверьте Excel файл и повторите попытку.\n"
                                  f"\nНайденные столбцы:\n {list_idx}\n"
                                  f"Не забудьте сохранить файл после редактирования!"
                )
                return

            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                system = str(row[system_idx] or "").strip()
                track = str(row[track_idx] or "").strip()
                cable = str(row[cable_idx] or "").strip()
                length_val = str(row[length_idx] or "").strip()
                try:
                    qty = int(row[quantity_idx])
                except:
                    qty = 1
                for _ in range(qty):
                    data.append(
                            {
                                    "system": system,
                                    "track": track,
                                    "cable": cable,
                                    "length": length_val
                            }
                    )

            output_path = os.path.join(output_dir, "cable_labels.pdf")
            c = canvas.Canvas(output_path, pagesize=A4)
            c.setFont("Times-Bold", 12)

            index = 0
            total_sides = len(data) * 2
            self.progress["maximum"] = total_sides
            self.progress["value"] = 0

            while index < len(data):
                # Лицевая сторона
                self.draw_page(c, data, index, side='front')
                c.showPage()
                self.progress["value"] += 1

                # Обратная сторона
                self.draw_page(c, data, index, side='back')
                if index + MAX_COLS * MAX_ROWS < len(data):
                    c.showPage()
                self.progress["value"] += 1

                index += MAX_COLS * MAX_ROWS

            c.save()
            messagebox.showinfo("Готово", f"PDF сохранён:\n{output_path}")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка: {str(e)}")

    def draw_page(self, c, data, start_index, side):
        """
        Рисует одну страницу с этикетками.
        :param c: объект canvas (PDF)
        :param data: список данных
        :param start_index: с какого элемента начинать
        :param side: 'front' или 'back'
        """
        col_step = TRIANGLE_BASE / 2
        x_centers_original = [45 * mm, 75 * mm, 105 * mm, 135 * mm, 165 * mm]
        Y_START = 76.5 * mm

        # Компенсация принтера
        shift_x = PRINTER_OFFSET_X * mm if side == 'back' else 0

        # Для обратной стороны — отзеркаливаем X координаты
        count = 0
        for i in range(start_index, min(start_index + MAX_COLS * MAX_ROWS, len(data))):
            item = data[i]
            col = count % MAX_COLS
            row = count // MAX_COLS

            if row >= MAX_ROWS:
                break

            # Базовая координата X (для лицевой стороны)
            center_x_base = x_centers_original[col]

            # Для обратной стороны — отзеркаливаем относительно центра листа
            if side == 'back':
                center_x = PAGE_WIDTH - center_x_base + shift_x
            else:
                center_x = center_x_base + shift_x

            y_base = Y_START + row * TRIANGLE_HEIGHT
            is_upside_down = col % 2 == 1

            if side == 'front':
                main_text = item["system"]
                sub_text = item["track"]
                main_font = FONT_SYSTEM
                sub_font = FONT_TRACK
            else:
                main_text = item["cable"]
                raw_length = str(item["length"]).strip()
                sub_text = f"L={raw_length} м" if raw_length.replace('.', '').isdigit() else raw_length
                main_font = FONT_CABLE
                sub_font = FONT_LENGTH

            self.draw_triangle(
                    c, center_x, y_base, is_upside_down, main_text, sub_text,
                    main_font, sub_font, side
            )

            count += 1

    def draw_triangle(
            self, c, center_x, y_base, upside_down, main_text, sub_text,
            main_font_size, sub_font_size, side
    ):
        """
        Рисует один треугольник с текстом.
        :param c: canvas
        :param center_x: X центра основания
        :param y_base: Y основания (зависит от ориентации)
        :param upside_down: True если треугольник остриём вниз
        :param main_text: основной текст (system/cable)
        :param sub_text: подзаголовок (track/length)
        :param main_font_size: размер шрифта основного текста
        :param sub_font_size: размер шрифта подзаголовка
        :param side: 'front' или 'back'
        """
        # 🔍 Отладка ширины
        test_text = "ШЩЖДМФЩЮДП"
        try:
            w = pdfmetrics.stringWidth(test_text, "Times-Bold", 12)
            print(f"📏 stringWidth работает: '{test_text}' → {w:.1f} pt")
        except Exception as e:
            print(f"❌ stringWidth НЕ РАБОТАЕТ: {e}")
        if not main_text.strip() and not sub_text.strip():
            return

        base = TRIANGLE_BASE
        height = TRIANGLE_HEIGHT
        x_left = center_x - base / 2
        x_right = center_x + base / 2

        # Определяем вершины треугольника
        if upside_down:
            points = [(x_left, y_base), (x_right, y_base), (center_x, y_base - height)]
        else:
            points = [(x_left, y_base - height), (x_right, y_base - height), (center_x, y_base)]

        # Рисуем контур
        c.setLineWidth(5.0)
        c.setStrokeColorRGB(0, 0, 0)
        c.lines(
                [
                        (points[0][0], points[0][1], points[1][0], points[1][1]),
                        (points[1][0], points[1][1], points[2][0], points[2][1]),
                        (points[2][0], points[2][1], points[0][0], points[0][1])
                ]
        )

        # Относительные смещения текста от основания
        dy_main = height * 0.35  # Основной текст ближе к центру
        dy_sub = height * 0.1  # Подзаголовок у основания

        c.saveState()

        if upside_down:
            # Поворачиваем вокруг центра основания
            c.translate(center_x, y_base)
            c.rotate(180)
            c.translate(-center_x, -y_base)
            y_main = y_base + dy_main
            y_sub = y_base + dy_sub
        else:
            base_y = y_base - height
            y_main = base_y + dy_main
            y_sub = base_y + dy_sub

        # --- ОСНОВНОЙ ТЕКСТ (system или cable) ---
        fs = main_font_size
        lines = []

        if side == 'back':  # cable — разбиваем на 2 строки
            parts = split_cable_text(main_text)
            line1, line2 = parts[0], parts[1]
            lines = [line1, line2]

            # Уменьшаем, если вторая часть длинная
            if len(line2) >= 15:
                fs = 12
            elif len(main_text) >= 20:
                fs = 14
            else:
                fs = FONT_CABLE  # 16

        else:  # Лицевая сторона — system
            lines = [main_text]
            # Уменьшаем, если строка >= 8 символов
            if len(main_text) >= 8:
                fs = 16
            else:
                fs = FONT_SYSTEM  # 18

        # Ограничиваем минимальным размером
        if fs < MIN_FONT_SIZE:
            fs = MIN_FONT_SIZE

        # Позиции строк (ваши идеальные настройки)
        if side == 'back':
            y_upper = y_main - fs * 0.5
            y_lower = y_main + fs * 0.5
            y_positions = [y_lower, y_upper]  # как вы сказали — работает идеально
        else:
            y_positions = [y_main]

        # print(f"🔧 [DEBUG] side={side}, main_text='{main_text}', len={len(main_text)}, fs={fs}")

        c.setFont("Times-Bold", fs)
        for j, line in enumerate(lines):
            # Грубая оценка ширины: символ ≈ 0.6 * шрифт
            try:
                tw = pdfmetrics.stringWidth(line, "Times-Bold", fs)
            except:
                tw = len(line) * fs * 0.6
            x_pos = center_x - tw / 2
            y_pos = y_positions[j]
            c.drawString(x_pos, y_pos, line)

        # --- ПОДЗАГОЛОВОК (track или length) ---
        max_chars_per_line = 28  # Подобрано под 60 мм и font=14
        if side == 'front' and len(sub_text) == 18:
            track_font_size = 13.5
            max_chars_per_line = 33  # При меньшем шрифте — можно больше символов
        elif side == 'front' and len(sub_text) == 19:
            track_font_size = 13
            max_chars_per_line = 33  # При меньшем шрифте — можно больше символов
        elif side == 'front' and len(sub_text) >= 20:
            track_font_size = 11.5
            max_chars_per_line = 39 # При меньшем шрифте — можно больше символов
        else:
            track_font_size = sub_font_size  # 14

        # Разбивка на 2 строки по длине
        max_len = 30 if track_font_size > 12 else 38
        line1 = sub_text[:max_len].strip()
        line2 = sub_text[max_len:max_len * 2].strip()

        lines = []
        if line1:
            lines.append(line1)
        if line2:
            lines.append(line2)

        # Устанавливаем шрифт
        c.setFont("Times-Bold", track_font_size)

        line_height = track_font_size * 1.5

        for j, line in enumerate(lines):
            if not line.strip():
                continue

        # ⚡️ Реальная ширина через stringWidth
        try:
            tw = pdfmetrics.stringWidth(line, "Times-Bold", track_font_size)
            # print(f"📏 Точная ширина: '{line}' → {tw:.1f} pt")
        except:
            # Fallback: улучшенная оценка с коэффициентом для кириллицы
            # Коэффициент 0.65 вместо 0.55 — лучше для широких букв
            estimated_width_per_char = {
                            'Ш': 1.2, 'Щ': 1.2, 'Ж': 1.15, 'Д': 1.1, 'П': 1.05,
                            'А': 0.9, 'В': 0.95, 'Е': 0.9, 'К': 0.95, 'Х': 0.9
                    }
            total_width = 0
            for char in line.upper():
                total_width += estimated_width_per_char.get(char, 1.0)
            tw = total_width * track_font_size * 0.58

        x_pos = center_x - tw / 2
        y_pos = y_sub - j * line_height
        c.drawString(x_pos, y_pos, line)

        # -----------------------------------------------------------------
        # Для track на лицевой стороне — уменьшаем шрифт при длинной строке
        # track_font_size = 12 if (side == 'front' and len(sub_text) >= 20) else sub_font_size
        #
        # # Разбиваем текст на строки (максимум 2)
        # max_width = base * 0.9
        # lines_sub = []
        # words = sub_text.split()
        # line = ""
        # for word in words:
        #     test = f"{line} {word}".strip()
        #     try:
        #         w = pdfmetrics.stringWidth(test, "Times-Bold", track_font_size)
        #     except:
        #         w = len(test) * track_font_size * 0.6
        #     if w <= max_width:
        #         line = test
        #     else:
        #         if line:
        #             lines_sub.append(line)
        #         line = word
        # if line:
        #     lines_sub.append(line)
        # lines_sub = lines_sub[:2]
        #
        # c.setFont("Times-Bold", track_font_size)
        # line_height = track_font_size * 1.4
        # for j, line in enumerate(lines_sub):
        #     try:
        #         tw = pdfmetrics.stringWidth(line, "Times-Bold", track_font_size)
        #     except:
        #         tw = len(line) * track_font_size * 0.6
        #     c.drawString(center_x - tw / 2, y_sub - j * line_height, line)

        c.restoreState()


if __name__ == "__main__":
    root = tk.Tk()
    app = CableLabelApp(root)
    root.mainloop()